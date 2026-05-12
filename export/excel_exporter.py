from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from export.context import ExportContext
from export.excel_columns import EXCEL_COLUMNS, build_excel_rows
from export.excel_images import add_images_to_sheet, get_embed_image_paths

EMBED_IMAGE_HEADER_PREFIX = "图片预览"
WIDE_COLUMNS = {
    "帖子发送时间": 20,
    "帖子链接": 45,
    "原图链接": 80,
    "图片本地路径": 80,
    "帖子内容": 60,
    "热评1(按点赞)": 65,
    "热评2(按点赞)": 65,
    "热评3(按点赞)": 65,
}


def export_excel(ctx: ExportContext, path: Path | None = None) -> Path:
    target = path or ctx.run_dir / "weibo_posts.xlsx"
    workbook = build_excel_workbook(ctx)
    save_excel(workbook, target)
    return target


def export_posts_xlsx(
    posts: Iterable[dict[str, Any]],
    xlsx_path: Path,
    column_map: list[tuple[str, str]] | None = None,
) -> None:
    ctx = ExportContext(
        run_dir=xlsx_path.parent,
        selected_posts=list(posts),
        all_posts=[],
        config={},
        stats={},
    )
    workbook = build_excel_workbook(ctx, column_map=column_map)
    save_excel(workbook, xlsx_path)


def build_excel_workbook(
    ctx: ExportContext,
    column_map: list[tuple[str, str]] | None = None,
):
    rows = list(ctx.selected_posts)
    columns = column_map or EXCEL_COLUMNS
    workbook = Workbook()
    write_posts_sheet(workbook, ctx, rows=rows, column_map=columns)
    return workbook


def write_posts_sheet(
    workbook,
    ctx: ExportContext,
    rows: list[dict[str, Any]] | None = None,
    column_map: list[tuple[str, str]] | None = None,
) -> None:
    posts = list(rows if rows is not None else ctx.selected_posts)
    columns = column_map or EXCEL_COLUMNS
    ws = cast(Worksheet, workbook.active)
    ws.title = "帖子统计"

    base_headers = [cn for _, cn in columns]
    max_embed_images = max((len(get_embed_image_paths(row)) for row in posts), default=0)
    image_headers = [f"{EMBED_IMAGE_HEADER_PREFIX}{i}" for i in range(1, max_embed_images + 1)]
    headers = base_headers + image_headers
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_idx)].width = 26 if header.startswith(EMBED_IMAGE_HEADER_PREFIX) else WIDE_COLUMNS.get(header, 14)

    for row in build_excel_rows(posts, columns):
        ws.append([row[h] for h in base_headers] + [""] * max_embed_images)

    wrap_cols = {"帖子内容", "原图链接", "图片本地路径", "热评1(按点赞)", "热评2(按点赞)", "热评3(按点赞)"}
    for col_idx, header in enumerate(headers, start=1):
        if header in wrap_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    if max_embed_images:
        first_image_col = len(base_headers) + 1
        for row_idx, post in enumerate(posts, start=2):
            try:
                add_images_to_sheet(ws, post, row_idx, first_image_col, max_embed_images)
            except Exception as err:
                ctx.warnings.append(f"Excel 图片嵌入失败：{type(err).__name__}")


def save_excel(workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
